"""
标准化处理通用工具
"""

from __future__ import annotations

import os
import re
from typing import List, Optional, Tuple, Dict, Any
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def _read_markdown_content_lines(file_path: str) -> List[str]:
    """读取markdown文件中首尾三个反引号之间的内容行。"""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    with open(file_path, 'r', encoding='utf-8') as f:
        all_lines = f.readlines()

    content_start = -1
    for i, line in enumerate(all_lines):
        if line.strip() == "```" and i > 0:
            content_start = i + 1
            break

    if content_start == -1:
        raise ValueError("无法找到markdown文本内容")

    content_end = len(all_lines)
    for i in range(len(all_lines) - 1, 0, -1):
        if all_lines[i].strip() == "```":
            content_end = i
            break

    return all_lines[content_start:content_end]


def chunk_file_by_lines(file_path: str, lines_per_chunk: int) -> List[Tuple[List[str], Optional[List[str]]]]:
    """按行数切分markdown内容，返回 (chunk1, chunk2) 列表。"""
    content_lines = _read_markdown_content_lines(file_path)

    chunks: List[Tuple[List[str], Optional[List[str]]]] = []
    for i in range(0, len(content_lines), lines_per_chunk):
        chunk1_start = i
        chunk1_end = min(i + lines_per_chunk, len(content_lines))
        chunk1 = content_lines[chunk1_start:chunk1_end]

        chunk2: Optional[List[str]] = None
        if chunk1_end < len(content_lines):
            chunk2_end = min(chunk1_end + lines_per_chunk, len(content_lines))
            chunk2 = content_lines[chunk1_end:chunk2_end]

        chunks.append((chunk1, chunk2))

    return chunks


def call_openai_with_retries(client: Any, model: str, prompt: str, max_retries: int, temperature: float = 0.1) -> Optional[str]:
    """带重试的OpenAI对话调用。"""
    for attempt in range(max_retries):
        try:
            response = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                temperature=temperature,
            )
            return response.choices[0].message.content
        except Exception as exc:  # noqa: BLE001 - 打印异常信息
            print(f"API调用失败 (尝试 {attempt + 1}/{max_retries}): {exc}")
            if attempt == max_retries - 1:
                return None
    return None


def split_questions_by_separator(ai_response: str, separator: str = "=== 题目分隔符 ===") -> List[str]:
    """将标准化文本按分隔符拆分为题目；若无分隔符，回退按 '### 试题 ' 块拆分。"""
    if not ai_response:
        return []

    # 首选：按照显式分隔符拆分
    if separator in ai_response:
        parts = ai_response.split(separator)
        results: List[str] = []
        for part in parts:
            text = part.strip()
            if text:
                results.append(text)
        if results:
            return results

    # 回退：根据 '### 试题 ' 标题切分
    # 捕获从每个 '### 试题 ' 开头到下一个 '### 试题 ' 之前的内容
    pattern = r"(### 试题\s+\d+[\s\S]*?)(?=\n### 试题\s+\d+|$)"
    matches = re.findall(pattern, ai_response)
    fallback_results: List[str] = []
    for block in matches:
        text = block.strip()
        if text:
            fallback_results.append(text)
    return fallback_results


def save_markdown_chunk_result(chunk_index: int, questions: List[str], output_dir: str, question_type_name: str) -> None:
    """保存某个chunk的标准化结果为markdown文件。"""
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    chunk_file = os.path.join(output_dir, f"standardized_chunk_{chunk_index:03d}.md")
    with open(chunk_file, 'w', encoding='utf-8') as f:
        f.write(f"# {question_type_name} - Chunk {chunk_index}\n\n")
        f.write(f"## 标准化时间\n{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write(f"## 题目数量\n{len(questions)}\n\n")
        f.write("## 标准化题目\n\n")
        for i, question in enumerate(questions, 1):
            f.write(f"### 题目 {i}\n\n")
            f.write(f"```\n{question}\n```\n\n")

    print(f"✅ Chunk {chunk_index}: 保存 {len(questions)} 道题目到 {chunk_file}")


def save_original_chunk_file(chunk_index: int, chunk1: List[str], output_dir: str) -> None:
    """保存原始chunk内容到markdown文件。"""
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    chunk_file = os.path.join(output_dir, f"original_chunk_{chunk_index:03d}.md")
    with open(chunk_file, 'w', encoding='utf-8') as f:
        f.write(f"# 原始 Chunk {chunk_index}\n\n")
        f.write(f"## 生成时间\n{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write("## 分块内容\n\n")
        f.write(''.join(chunk1))

    print(f"💾 原始 Chunk {chunk_index} 已保存到: {chunk_file}")


def list_standardized_chunk_files(standardized_dir: str) -> List[str]:
    """列出标准化目录下所有标准化chunk文件（按序）。"""
    files = [
        os.path.join(standardized_dir, f)
        for f in os.listdir(standardized_dir)
        if f.startswith('standardized_chunk_') and f.endswith('.md')
    ]
    files.sort()
    return files


def extract_codeblocks_from_markdown(content: str) -> List[str]:
    """从markdown文本中提取所有三反引号包裹的代码块内容。"""
    return re.findall(r"```\n(.*?)\n```", content, re.DOTALL)


def write_excel(headers: List[str], rows: List[List[Any]], sheet_title: str, output_path: str) -> None:
    """创建Excel并写入表头与行数据，自动列宽。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_index, row_values in enumerate(rows, 2):
        for col_index, value in enumerate(row_values, 1):
            ws.cell(row=row_index, column=col_index, value=value)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"✅ Excel文件已保存: {output_path}")
    print(f"📊 共写入 {len(rows)} 行")


