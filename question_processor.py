"""
题库处理器 - 主要模块
用于从PDF文档中提取题目并转换为结构化的Excel文件
"""

import fitz  # PyMuPDF
import openpyxl
import json
import re
import os
import tempfile
import datetime
from typing import List, Dict, Optional, Tuple
from tqdm import tqdm


class QuestionProcessor:
    """题库处理器主类（仅保留PDF解析与题型拆分工具方法，AI流程已移除）"""
    
    # 定义题型映射
    QUESTION_TYPE_MAPPING = {
        "单选题": {"name": "single_choice", "pattern": r"二、单选题：.*[（(](\d+).*题[）)]"},
        "多选题": {"name": "multiple_choice", "pattern": r"三、多选题：.*[（(](\d+).*题[）)]"},
        "判断题": {"name": "judgment", "pattern": r"四、判断题：.*[（(](\d+).*题[）)]"},
        "简答题": {"name": "short_answer", "pattern": r"五、简答题.*[（(](\d+).*题[）)]"},
        "论述题": {"name": "essay", "pattern": r"六、论述题.*[（(](\d+).*题[）)]"},
        "案例分析题": {"name": "case_analysis", "pattern": r"七、案例分析题：.*[（(](\d+).*题[）)]"}
    }
    
    def __init__(self):
        """初始化（不再依赖 Gemini）"""
        pass
    
    def extract_text_from_pdf(self, pdf_path: str) -> str:
        """从PDF文件中提取所有文本"""
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"PDF文件不存在: {pdf_path}")
        
        try:
            doc = fitz.open(pdf_path)
            full_text = ""
            
            for page in doc:
                full_text += page.get_text() + "\n"
            
            doc.close()
            return full_text
            
        except Exception as e:
            raise Exception(f"PDF处理错误: {e}") from e
    
    def split_text_by_question_types(self, text: str, output_dir: str) -> Dict[str, Dict]:
        """
        按题型拆分文本并保存到子文件
        
        Args:
            text: 完整的PDF文本
            output_dir: 输出目录
            
        Returns:
            Dict: 包含各题型信息的字典 {题型名: {"file_path": 路径, "text": 文本, "count": 预期题数}}
        """
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        lines = text.split('\n')
        question_type_sections = {}
        
        # 找到各题型的起始位置
        section_positions = []
        for i, line in enumerate(lines):
            line = line.strip()
            for type_name, type_info in self.QUESTION_TYPE_MAPPING.items():
                if re.search(type_info["pattern"], line):
                    # 提取预期题数
                    count_match = re.search(r'[（(](\d+).*题[）)]', line)
                    expected_count = int(count_match.group(1)) if count_match else 0
                    
                    section_positions.append({
                        "type_name": type_name,
                        "type_info": type_info,
                        "line_index": i,
                        "expected_count": expected_count,
                        "section_line": line
                    })
                    print(f"找到题型: {type_name}，预期题数: {expected_count}")
                    break
        
        # 按行号排序
        section_positions.sort(key=lambda x: x["line_index"])
        
        # 提取各题型的文本内容
        for i, section in enumerate(section_positions):
            type_name = section["type_name"]
            type_info = section["type_info"]
            start_line = section["line_index"]
            expected_count = section["expected_count"]
            
            # 确定结束位置（下一个题型的开始或文档结束）
            if i + 1 < len(section_positions):
                end_line = section_positions[i + 1]["line_index"]
            else:
                end_line = len(lines)
            
            # 提取该题型的文本
            section_text = '\n'.join(lines[start_line:end_line])
            
            # 保存到markdown文件
            filename = f"{type_info['name']}.md"
            file_path = os.path.join(output_dir, filename)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(f"# {type_name} ({expected_count}题)\n\n")
                f.write(f"## 提取时间\n{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                f.write(f"## 预期题数\n{expected_count}\n\n")
                f.write(f"## 原始文本\n\n```\n{section_text}\n```\n")
            
            question_type_sections[type_name] = {
                "file_path": file_path,
                "text": section_text,
                "expected_count": expected_count,
                "type_info": type_info
            }
            
            print(f"✅ 保存 {type_name} 到: {file_path}")
        
        return question_type_sections
    
    def split_text_into_questions(self, text: str) -> List[str]:
        """将长文本分割成独立的题目块 - 改进版本支持多个section的重复编号"""
        if not text or not text.strip():
            return []
        
        # 使用正则表达式找到所有题目编号位置
        question_pattern = r'\n\s*(\d+)[.．、]\s*[（(]?'
        
        # 分割文本，但保留分割符
        parts = re.split(question_pattern, text)
        
        if len(parts) <= 1:
            return []
        
        questions = []
        
        # 重新组装题目，将编号和内容合并
        for i in range(1, len(parts), 2):
            if i + 1 < len(parts):
                question_num = parts[i]
                question_content = parts[i + 1]
                
                # 重新构建完整的题目文本
                full_question = f"{question_num}. {question_content.strip()}"
                
                # 过滤掉过短的文本块（可能是章节标题等）
                if len(full_question) > 20:
                    questions.append(full_question)
        
        # 旧的非AI补救逻辑已移除
        
        return questions
    
    def process_single_question_type(self, question_type_file: str, output_dir: str = None) -> List[Dict]:
        """
        处理单个题型文件
        
        Args:
            question_type_file: 题型文件路径（markdown格式）
            output_dir: 输出目录，如果不指定则使用默认目录
            
        Returns:
            List[Dict]: 处理后的题目数据列表
        """
        if not os.path.exists(question_type_file):
            raise FileNotFoundError(f"题型文件不存在: {question_type_file}")
        
        # 读取题型文件
        with open(question_type_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 提取题型名称和原始文本
        type_name = ""
        if "# " in content:
            type_line = content.split('\n')[0]
            type_name = type_line.replace('# ', '').split(' (')[0]
        
        # 提取原始文本部分
        text_start = content.find("```\n") + 4
        text_end = content.rfind("\n```")
        if text_start > 3 and text_end > text_start:
            section_text = content[text_start:text_end]
        else:
            raise ValueError("无法从题型文件中提取原始文本")
        
        print(f"开始处理题型: {type_name}")
        
        # 设置输出目录
        if output_dir is None:
            base_dir = os.path.dirname(question_type_file)
            output_dir = os.path.join(base_dir, f"{type_name}_questions")
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        print(f"题目输出目录: {output_dir}")
        
        # 分割成独立题目
        questions = self.split_text_into_questions(section_text)
        print(f"从 {type_name} 中提取到 {len(questions)} 道题目")
        
        # 保存每道题目到markdown文件
        saved_count = 0
        for i, question_text in enumerate(questions, 1):
            markdown_path = self.save_question_to_markdown(i, question_text, output_dir)
            if markdown_path:
                saved_count += 1
        
        print(f"已保存 {saved_count} 道题目的markdown文件到: {output_dir}")
        
        # 使用AI处理每道题目
        print(f"正在使用AI处理 {type_name} 题目...")
        processed_questions = []
        
        for i, question_text in enumerate(tqdm(questions, desc=f"处理{type_name}")):
            try:
                structured_data = self.get_structured_data_from_ai(question_text)
                if structured_data:
                    # 添加题目序号和类型信息
                    structured_data['source_index'] = i + 1
                    structured_data['source_type'] = type_name
                    structured_data['markdown_file'] = f"question_{i+1:04d}.md"
                    processed_questions.append(structured_data)
                else:
                    print(f"警告：第{i+1}道{type_name}题目处理失败，跳过")
            except Exception as e:
                print(f"警告：第{i+1}道{type_name}题目处理出错: {e}")
                continue
        
        print(f"AI处理完成，成功处理 {len(processed_questions)} 道{type_name}题目")
        
        # 保存到对应的Excel文件
        excel_path = os.path.join(output_dir, f"{type_name}_processed.xlsx")
        self.save_to_excel(processed_questions, excel_path)
        print(f"结果已保存到: {excel_path}")
        
        return processed_questions
    
    def split_questions_only(self, question_type_file: str, output_dir: str = None) -> int:
        """
        仅将题型文件拆分为独立的题目文件，不进行AI处理
        
        Args:
            question_type_file: 题型文件路径（markdown格式）
            output_dir: 输出目录，如果不指定则使用默认目录
            
        Returns:
            int: 拆分出的题目数量
        """
        if not os.path.exists(question_type_file):
            raise FileNotFoundError(f"题型文件不存在: {question_type_file}")
        
        # 读取题型文件
        with open(question_type_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 提取题型名称和预期题数
        type_name = ""
        expected_count = 0
        if "# " in content:
            type_line = content.split('\n')[0]
            type_name = type_line.replace('# ', '').split(' (')[0]
            # 提取预期题数
            if '(' in type_line and '题)' in type_line:
                count_part = type_line.split('(')[1].split('题)')[0]
                try:
                    expected_count = int(count_part)
                except ValueError:
                    expected_count = 0
        
        # 提取原始文本部分
        text_start = content.find("```\n") + 4
        text_end = content.rfind("\n```")
        if text_start > 3 and text_end > text_start:
            section_text = content[text_start:text_end]
        else:
            raise ValueError("无法从题型文件中提取原始文本")
        
        print(f"开始拆分题型: {type_name}")
        print(f"预期题数: {expected_count}")
        
        # 设置输出目录
        if output_dir is None:
            base_dir = os.path.dirname(question_type_file)
            output_dir = os.path.join(base_dir, f"{type_name}_questions")
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        print(f"题目输出目录: {output_dir}")
        
        # 分割成独立题目
        questions = self.split_text_into_questions(section_text)
        actual_count = len(questions)
        print(f"从 {type_name} 中提取到 {actual_count} 道题目 (预期: {expected_count})")
        
        # 保存每道题目到markdown文件
        saved_count = 0
        for i, question_text in enumerate(questions, 1):
            markdown_path = self.save_question_to_markdown(i, question_text, output_dir)
            if markdown_path:
                saved_count += 1
        
        print(f"✅ 已保存 {saved_count} 道题目的markdown文件到: {output_dir}")
        
        # 生成拆分统计信息
        stats_file = os.path.join(output_dir, "split_stats.md")
        with open(stats_file, 'w', encoding='utf-8') as f:
            f.write(f"# {type_name} 拆分统计\n\n")
            f.write(f"## 基本信息\n")
            f.write(f"- 题型: {type_name}\n")
            f.write(f"- 预期题数: {expected_count}\n")
            f.write(f"- 实际提取: {actual_count} 道题目\n")
            f.write(f"- 保存成功: {saved_count} 个文件\n")
            f.write(f"- 拆分时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write(f"## 提取率\n")
            if expected_count > 0:
                extraction_rate = (actual_count / expected_count) * 100
                f.write(f"- 提取率: {extraction_rate:.1f}% ({actual_count}/{expected_count})\n")
            else:
                f.write(f"- 提取率: 无法计算（预期题数为0）\n")
            
            if actual_count != expected_count:
                f.write(f"\n## ⚠️ 注意事项\n")
                if actual_count < expected_count:
                    f.write(f"- 实际提取数量少于预期，可能存在题目格式识别问题\n")
                    f.write(f"- 建议检查原始文本格式，优化题目分割算法\n")
                else:
                    f.write(f"- 实际提取数量多于预期，可能存在重复或误识别\n")
                    f.write(f"- 建议抽样检查题目内容的正确性\n")
            
            f.write(f"\n## 文件列表\n")
            for i in range(1, saved_count + 1):
                f.write(f"- question_{i:04d}.md\n")
        
        print(f"📊 统计信息已保存到: {stats_file}")
        return actual_count
    
    def save_question_to_markdown(self, question_index: int, question_text: str, temp_dir: str) -> str:
        """将单个题目保存为markdown文件用于追溯"""
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
        
        # 创建文件名，使用4位数补零
        filename = f"question_{question_index:04d}.md"
        filepath = os.path.join(temp_dir, filename)
        
        # 准备markdown内容
        markdown_content = f"""# 题目 {question_index}

## 原始文本

```
{question_text}
```

## 提取时间
{__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## 文本长度
{len(question_text)} 字符

---
"""
        
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(markdown_content)
            return filepath
        except Exception as e:
            print(f"警告：无法保存题目{question_index}的markdown文件: {e}")
            return ""
    
    def get_structured_data_from_ai(self, question_text: str) -> Optional[Dict]:
        """调用AI从单个题目文本中提取结构化信息"""
        if not question_text or not question_text.strip():
            return None
        system_prompt = """
你是一个专业的题目解析助手。请从给定的题目文本中提取结构化信息，并以JSON格式返回。

返回格式要求:
{
    "question_type": "题目类型（单选/多选/判断/填空/简答）",
    "difficulty": "难度标记（如果有）",
    "question_stem": "题目主干内容",
    "options": ["选项列表（如果有）"],
    "answer": "参考答案",
    "explanation": "解析内容（如果有）"
}

注意:
1. 严格按照上述JSON格式返回
2. 如果没有选项，options字段返回空数组[]
3. 如果没有解析，explanation字段返回空字符串""
4. 如果没有难度标记，difficulty字段返回空字符串""
5. 准确识别题目类型：单选、多选、判断、填空、简答
"""

        user_prompt = f"请解析以下题目：\n{question_text}"
        
        try:
            response = self.model.generate_content(
                f"{system_prompt}\n\n{user_prompt}"
            )
            
            # 安全地获取响应文本
            content = None
            if hasattr(response, 'text') and response.text:
                content = response.text
            elif hasattr(response, 'candidates') and response.candidates:
                candidate = response.candidates[0]
                if hasattr(candidate, 'content') and candidate.content.parts:
                    content = candidate.content.parts[0].text
            
            if not content:
                return None
                
            result = json.loads(content)
            
            # 验证必需字段
            required_fields = ["question_type", "difficulty", "question_stem", "options", "answer", "explanation"]
            for field in required_fields:
                if field not in result:
                    result[field] = "" if field != "options" else []
            
            return result
            
        except json.JSONDecodeError:
            # JSON解析失败
            return None
        except Exception as e:
            # API调用或其他错误
            return None
    
    def save_to_excel(self, questions_data: List[Dict], output_path: str) -> None:
        """将结构化数据保存到Excel文件"""
        if not questions_data and not os.path.exists(output_path):
            # 如果没有数据且不是基于现有模板，创建空的Excel文件
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "题库数据"
            
            # 设置表头
            headers = ["序号", "题型", "难度", "题目", "选项A", "选项B", "选项C", "选项D", "选项E", "答案", "解析"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            try:
                wb.save(output_path)
            except Exception as e:
                raise Exception(f"Excel文件保存失败: {e}") from e
            return
        
        # 如果输出文件已存在，加载现有文件（作为模板）
        if os.path.exists(output_path):
            try:
                wb = openpyxl.load_workbook(output_path)
            except Exception as e:
                # 如果加载失败，创建新的工作簿
                wb = openpyxl.Workbook()
        else:
            wb = openpyxl.Workbook()
        
        ws = wb.active
        if ws.title == "Sheet":
            ws.title = "题库数据"
        
        # 设置表头（如果是新建的工作簿）
        if ws.max_row == 1 and not ws.cell(row=1, column=1).value:
            headers = ["序号", "题型", "难度", "题目", "选项A", "选项B", "选项C", "选项D", "选项E", "答案", "解析"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
        
        # 找到数据开始行（表头后的第一个空行）
        start_row = ws.max_row + 1 if ws.max_row > 1 else 2
        
        # 逐行写入题目数据
        for idx, question in enumerate(questions_data):
            row = start_row + idx
            
            # 序号
            ws.cell(row=row, column=1, value=idx + 1)
            
            # 题型
            ws.cell(row=row, column=2, value=question.get("question_type", ""))
            
            # 难度
            ws.cell(row=row, column=3, value=question.get("difficulty", ""))
            
            # 题目
            ws.cell(row=row, column=4, value=question.get("question_stem", ""))
            
            # 选项（最多支持5个选项A-E）
            options = question.get("options", [])
            for i in range(5):  # 选项A-E对应列5-9
                if i < len(options):
                    ws.cell(row=row, column=5+i, value=options[i])
                else:
                    ws.cell(row=row, column=5+i, value="")
            
            # 答案
            ws.cell(row=row, column=10, value=question.get("answer", ""))
            
            # 解析
            ws.cell(row=row, column=11, value=question.get("explanation", ""))
        
        # 保存文件
        try:
            wb.save(output_path)
        except Exception as e:
            raise Exception(f"Excel文件保存失败: {e}") from e
    
    def process_questions(self, pdf_path: str, output_path: str, step: str = "full") -> None:
        """
        完整的题目处理流程，支持分步骤执行
        
        Args:
            pdf_path: PDF文件路径
            output_path: 输出文件路径
            step: 处理步骤 - "split", "split-questions", "process", "full"
        """
        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
        work_dir = f"question_processing_{base_name}"
        
        if not os.path.exists(work_dir):
            os.makedirs(work_dir)
        
        print(f"开始处理PDF文件: {pdf_path}")
        print(f"工作目录: {work_dir}")
        print(f"执行步骤: {step}")
        
        if step in ["split", "full"]:
            # 步骤1: 提取PDF文本并按题型拆分
            print("\n" + "="*50)
            print("步骤1: 提取PDF文本并按题型拆分")
            print("="*50)
            
            full_text = self.extract_text_from_pdf(pdf_path)
            print(f"提取完成，共{len(full_text)}字符")
            
            # 按题型拆分
            type_sections_dir = os.path.join(work_dir, "question_types")
            question_type_sections = self.split_text_by_question_types(full_text, type_sections_dir)
            
            print(f"\n✅ 按题型拆分完成，共找到 {len(question_type_sections)} 个题型")
            for type_name, info in question_type_sections.items():
                print(f"  📋 {type_name}: {info['expected_count']} 题 -> {info['file_path']}")
        
        if step in ["split-questions", "full"]:
            # 步骤2: 将各题型文件拆分为独立题目文件
            print("\n" + "="*50)
            print("步骤2: 将题型文件拆分为独立题目")
            print("="*50)
            
            type_sections_dir = os.path.join(work_dir, "question_types")
            if not os.path.exists(type_sections_dir):
                raise FileNotFoundError("题型分离文件不存在，请先执行 step='split'")
            
            # 查找所有题型文件
            type_files = []
            for filename in os.listdir(type_sections_dir):
                if filename.endswith('.md'):
                    type_files.append(os.path.join(type_sections_dir, filename))
            
            if not type_files:
                raise FileNotFoundError("未找到题型文件，请先执行 step='split'")
            
            total_questions = 0
            split_summary = []
            
            # 拆分每个题型文件
            for type_file in sorted(type_files):
                try:
                    question_count = self.split_questions_only(type_file)
                    total_questions += question_count
                    
                    # 记录拆分信息
                    filename = os.path.basename(type_file)
                    type_name = filename.replace('.md', '').replace('_', ' ').title()
                    split_summary.append({
                        'file': filename,
                        'type': type_name,
                        'count': question_count
                    })
                    
                except Exception as e:
                    print(f"❌ 拆分题型文件 {type_file} 时出错: {e}")
                    continue
            
            print(f"\n✅ 所有题型拆分完成，共拆分出 {total_questions} 道独立题目")
            
            # 生成总体统计报告
            summary_file = os.path.join(work_dir, "split_summary.md")
            with open(summary_file, 'w', encoding='utf-8') as f:
                f.write(f"# 题目拆分总结报告\n\n")
                f.write(f"## 总体信息\n")
                f.write(f"- 处理时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"- 总题目数: {total_questions} 道\n")
                f.write(f"- 题型数量: {len(split_summary)} 个\n\n")
                f.write(f"## 各题型详情\n")
                
                for info in split_summary:
                    f.write(f"### {info['type']}\n")
                    f.write(f"- 文件: {info['file']}\n")
                    f.write(f"- 题目数: {info['count']} 道\n")
                    f.write(f"- 目录: {info['type'].lower().replace(' ', '_')}_questions/\n\n")
                
                f.write(f"## 目录结构\n")
                f.write(f"```\n")
                f.write(f"{work_dir}/\n")
                f.write(f"├── question_types/          # 题型分类文件\n")
                for info in split_summary:
                    f.write(f"├── {info['type'].lower().replace(' ', '_')}_questions/  # {info['count']}个题目文件\n")
                f.write(f"└── split_summary.md         # 本报告\n")
                f.write(f"```\n")
            
            print(f"📊 总结报告已保存到: {summary_file}")
            print(f"\n📁 各题型题目详情:")
            for info in split_summary:
                questions_dir = os.path.join(work_dir, f"{info['type'].lower().replace(' ', '_')}_questions")
                print(f"  📂 {info['type']}: {info['count']} 道题目 -> {questions_dir}")
        
        if step in ["process", "full"]:
            # 步骤3: AI处理各题型（原有的处理逻辑）
            print("\n" + "="*50)
            print("步骤3: AI处理各题型题目")
            print("="*50)
            
            # 这里保持原有的AI处理逻辑
            # 查找已经拆分的题目文件并进行AI处理
            print("⏳ AI处理功能将在下一步骤中实现...")
            print("💡 当前步骤已完成题目拆分，可以查看各题型的独立题目文件")
        
        print(f"\n🎉 步骤 '{step}' 处理完成！工作目录: {work_dir}")
        
        if step == "split-questions":
            print(f"\n📝 使用说明：")
            print("1. 查看生成的题目文件，验证拆分是否正确")
            print("2. 检查 split_summary.md 文件了解拆分统计")
            print("3. 运行 'python main.py --step process' 继续AI处理")
            print("4. 或者查看各题型目录中的 split_stats.md 了解详细信息")
    
    def _print_directory_tree(self, directory: str, indent: str = "") -> None:
        """打印目录树结构"""
        try:
            items = sorted(os.listdir(directory))
            for i, item in enumerate(items):
                path = os.path.join(directory, item)
                is_last = i == len(items) - 1
                current_indent = "└── " if is_last else "├── "
                
                if os.path.isdir(path):
                    print(f"{indent}{current_indent}{item}/")
                    next_indent = indent + ("    " if is_last else "│   ")
                    # 只显示前几个子项，避免输出过长
                    sub_items = sorted(os.listdir(path))[:5]
                    for j, sub_item in enumerate(sub_items):
                        sub_is_last = j == len(sub_items) - 1
                        sub_current_indent = "└── " if sub_is_last else "├── "
                        print(f"{next_indent}{sub_current_indent}{sub_item}")
                    if len(os.listdir(path)) > 5:
                        print(f"{next_indent}└── ... (还有{len(os.listdir(path)) - 5}个文件)")
                else:
                    print(f"{indent}{current_indent}{item}")
        except PermissionError:
            print(f"{indent}└── [权限不足]")